"""
Exporta todas as tabelas do banco de dados para um arquivo Excel.
Uma aba por tabela.

Uso:
    pip install openpyxl
    python export_db.py
    python export_db.py --db data/database.db --out exportacao.xlsx
"""

import sqlite3
import argparse
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instale openpyxl primeiro:  pip install openpyxl")
    raise SystemExit(1)


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill("solid", fgColor="D6E4F0")


def export(db_path: str, out_path: str) -> None:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
    tables = [r[0] for r in cur.fetchall() if not r[0].startswith("sqlite_")]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for table in tables:
        cur.execute(f"SELECT * FROM [{table}]")
        rows = cur.fetchall()
        columns = [d[0] for d in cur.description] if cur.description else []

        ws = wb.create_sheet(title=table[:31])  # Excel max 31 chars

        # Header row
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx % 2 == 0:
                    cell.fill = ALT_FILL

        # Auto-fit column widths (capped at 60)
        for col_idx, col_name in enumerate(columns, start=1):
            max_len = len(col_name)
            for row in rows:
                cell_val = str(row[col_idx - 1]) if row[col_idx - 1] is not None else ""
                max_len = max(max_len, len(cell_val))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

        # Freeze header row
        ws.freeze_panes = "A2"

        print(f"  {table}: {len(rows)} linha(s), {len(columns)} coluna(s)")

    conn.close()
    wb.save(out_path)
    print(f"\nArquivo salvo em: {out_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Exporta o banco SQLite para Excel")
    parser.add_argument("--db", default="data/database.db", help="Caminho do arquivo .db")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    parser.add_argument("--out", default=f"exportacao_{timestamp}.xlsx", help="Arquivo de saída")
    args = parser.parse_args()

    if not Path(args.db).exists():
        print(f"Banco não encontrado: {args.db}")
        raise SystemExit(1)

    print(f"Exportando '{args.db}'...")
    export(args.db, args.out)


if __name__ == "__main__":
    main()
