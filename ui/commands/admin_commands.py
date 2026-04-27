# -*- coding: utf-8 -*-
import io
import logging
from datetime import datetime

import discord
from discord.ext import commands

from core.config import IMAGE_CHANNEL_ID, ADMIN_IDS
from core.db.audit_repo import log_action
from core.db.lobby_repo import get_image_channel, set_image_channel, clear_image_channel
from core.db.match_repo import find_unregistered_match_players, diagnose_and_fix_kda_data, get_ranking_from_matches, fix_malformed_durations, fix_match_id_sequence, renumber_league_match
from core.db.player_repo import add_player_alias, remove_player_alias, get_player_aliases, get_player, upsert_player, get_all_player_aliases
from ui.commands.score_helpers import is_admin

audit_logger = logging.getLogger("Audit")

BOT_STATE: dict = {"enabled": True}
SEASON_STATE: dict = {"active": True}


def is_bot_enabled() -> bool:
    return BOT_STATE["enabled"]


def is_season_active() -> bool:
    return SEASON_STATE["active"]


def setup_admin_commands(bot: commands.Bot):
    logger = logging.getLogger("AdminCommands")
    logger.info("Carregando comandos administrativos...")

    @bot.command(name="registrarcanalimagem", aliases=["registrarcanalocr"])
    async def cmd_register_image_channel(ctx: commands.Context):
        if not is_admin(ctx.author.id):
            await ctx.message.delete()
            await ctx.send("❌ Apenas administradores.", delete_after=5)
            return

        if not ctx.guild:
            await ctx.message.delete()
            await ctx.send("❌ Este comando só pode ser usado em um servidor.", delete_after=8)
            return

        old_channel_id = get_image_channel(ctx.guild.id)
        set_image_channel(ctx.guild.id, ctx.channel.id)
        await ctx.message.delete()

        if old_channel_id and old_channel_id != ctx.channel.id:
            await ctx.send(
                f"✅ Canal de imagens atualizado: <#{old_channel_id}> → <#{ctx.channel.id}>",
                delete_after=20
            )
        else:
            await ctx.send(
                f"✅ Canal <#{ctx.channel.id}> registrado para leitura de imagens de partida.",
                delete_after=20
            )

    @bot.command(name="limparcanalimagem", aliases=["limparcanalocr"])
    async def cmd_clear_image_channel(ctx: commands.Context):
        if not is_admin(ctx.author.id):
            await ctx.message.delete()
            await ctx.send("❌ Apenas administradores.", delete_after=5)
            return

        if not ctx.guild:
            await ctx.message.delete()
            await ctx.send("❌ Este comando só pode ser usado em um servidor.", delete_after=8)
            return

        clear_image_channel(ctx.guild.id)
        await ctx.message.delete()
        await ctx.send(
            "✅ Canal de imagem OCR foi removido. Imagens só serão processadas em canais explicitamente configurados.",
            delete_after=15
        )

    @bot.command(name="canalimagem", aliases=["imagemcanal"])
    async def cmd_show_image_channel(ctx: commands.Context):
        if not is_admin(ctx.author.id):
            await ctx.message.delete()
            await ctx.send("❌ Apenas administradores.", delete_after=5)
            return

        if not ctx.guild:
            await ctx.message.delete()
            await ctx.send("❌ Este comando só pode ser usado em um servidor.", delete_after=8)
            return

        image_channel_id = get_image_channel(ctx.guild.id)
        if image_channel_id:
            await ctx.send(f"📌 Canal de imagem OCR registrado: <#{image_channel_id}>")
        elif IMAGE_CHANNEL_ID:
            await ctx.send(f"📌 Canal de imagem OCR definido via ENV: <#{IMAGE_CHANNEL_ID}>")
        else:
            await ctx.send("⚠️ Nenhum canal de imagem OCR registrado.")

    @bot.command(name="addalias", aliases=["aliasadd", "alias"])
    async def cmd_add_alias(ctx: commands.Context, member: discord.Member, *, alias: str):
        if not is_admin(ctx.author.id):
            await ctx.message.delete()
            await ctx.send("❌ Apenas administradores.", delete_after=5)
            return

        alias = alias.strip()
        if not alias:
            await ctx.send("❌ Informe o apelido após o membro.", delete_after=10)
            return

        add_player_alias(member.id, alias)
        log_action(
            ctx.author.id,
            ctx.author.display_name,
            "!addalias",
            f"discord_id={member.id} alias={alias}",
            affected_ids=[member.id]
        )

        await ctx.message.delete()
        await ctx.send(f"✅ Alias `{alias}` adicionado para {member.mention}.")

    @bot.command(name="cadastro")
    async def cmd_cadastro(ctx: commands.Context, member: discord.Member, *, nick: str):
        if not is_admin(ctx.author.id):
            await ctx.message.delete()
            await ctx.send("❌ Apenas administradores.", delete_after=5)
            return

        nick = nick.strip().strip('"\'')
        if not nick:
            await ctx.message.delete()
            await ctx.send("❌ Informe o nick.\n→ `!cadastro @usuario <nick>`", delete_after=600)
            return

        existing = get_player(member.id)
        if existing:
            upsert_player(member.id, existing["display_name"], existing["wins"], existing["losses"])
        else:
            upsert_player(member.id, member.display_name, 0, 0)

        add_player_alias(member.id, nick)
        log_action(
            ctx.author.id, ctx.author.display_name,
            "!cadastro",
            f"discord_id={member.id} nick={nick}",
            affected_ids=[member.id]
        )

        await ctx.message.delete()
        await ctx.send(f"✅ `{nick}` cadastrado como nick de {member.mention}.", delete_after=600)

    @bot.command(name="removealias", aliases=["delalias"])
    async def cmd_remove_alias(ctx: commands.Context, member: discord.Member, *, alias: str):
        if not is_admin(ctx.author.id):
            await ctx.message.delete()
            await ctx.send("❌ Apenas administradores.", delete_after=5)
            return

        alias = alias.strip()
        if not alias:
            await ctx.send("❌ Informe o alias a ser removido.", delete_after=10)
            return

        remove_player_alias(member.id, alias)
        log_action(
            ctx.author.id,
            ctx.author.display_name,
            "!removealias",
            f"discord_id={member.id} alias={alias}",
            affected_ids=[member.id]
        )

        await ctx.message.delete()
        await ctx.send(f"✅ Alias `{alias}` removido para {member.mention}.")

    @bot.command(name="aliases", aliases=["aliaslist"])
    async def cmd_alias_list(ctx: commands.Context, member: discord.Member):
        if not is_admin(ctx.author.id):
            await ctx.message.delete()
            await ctx.send("❌ Apenas administradores.", delete_after=5)
            return

        aliases = get_player_aliases(member.id)
        if not aliases:
            await ctx.send(f"⚠️ Nenhum alias registrado para {member.mention}.", delete_after=10)
            return

        alias_list = '\n'.join(f"- `{alias}`" for alias in aliases)
        await ctx.send(f"📝 Aliases para {member.mention}:\n{alias_list}")

    @bot.command(name="listaraliases", aliases=["listalias", "allaliases", "todosaliases"])
    async def cmd_listar_aliases(ctx: commands.Context):
        if not is_admin(ctx.author.id):
            await ctx.message.delete()
            await ctx.send("❌ Apenas administradores.", delete_after=5)
            return

        rows = get_all_player_aliases()
        if not rows:
            await ctx.send("⚠️ Nenhum alias cadastrado.")
            return

        lines = [
            f"{r['alias']:<25} → {r['display_name'] or ('<@' + str(r['discord_id']) + '>')}"
            for r in rows
        ]

        header = f"📋 **{len(lines)} alias(es) cadastrados:**"
        chunk_size = 1800
        text = "\n".join(lines)
        chunks = [text[i:i+chunk_size] for i in range(0, len(text), chunk_size)]

        await ctx.send(header)
        for chunk in chunks:
            await ctx.send(f"```\n{chunk}\n```")

    @bot.command(name="jogadoresfaltando", aliases=["missingplayers", "faltando"])
    async def cmd_jogadores_faltando(ctx: commands.Context):
        if not is_admin(ctx.author.id):
            await ctx.message.delete()
            await ctx.send("❌ Apenas administradores.", delete_after=5)
            return

        missing = find_unregistered_match_players()
        if not missing:
            await ctx.send("✅ Todos os jogadores das partidas importadas estão registrados.")
            return

        lines = [f"⚠️ **{len(missing)} jogador(es) em partidas mas sem registro na tabela `players`:**\n"]
        for m in missing:
            lines.append(
                f"• `{m['player_name']}` — discord_id: `{m['discord_id']}` — {m['partidas']} partida(s)\n"
                f"  → `!registrar <@{m['discord_id']}> 0 0` para registrar"
            )
        await ctx.send("\n".join(lines))

    @bot.command(name="fixkda", aliases=["corrigirkda"])
    async def cmd_fix_kda(ctx: commands.Context, confirm: str = ""):
        if not is_admin(ctx.author.id):
            await ctx.message.delete()
            await ctx.send("❌ Apenas administradores.", delete_after=5)
            return

        do_fix = confirm.lower() in {"sim", "yes", "ok", "fix"}
        result = diagnose_and_fix_kda_data(fix=do_fix)
        bad_rows = result["bad_rows"]

        if not bad_rows:
            await ctx.send("✅ Nenhum valor inválido de KDA encontrado nas partidas.")
            return

        lines = [f"⚠️ Encontrados **{len(bad_rows)}** registros com KDA inválido:"]
        for r in bad_rows[:20]:
            lines.append(
                f"• Partida {r['league_match_id']}, slot {r['slot']}, `{r['player_name']}` → "
                f"kills={r['kills']}({r['kills_type']}), deaths={r['deaths']}({r['deaths_type']}), assists={r['assists']}({r['assists_type']})"
            )
        if len(bad_rows) > 20:
            lines.append(f"... e mais {len(bad_rows) - 20} registros.")

        if do_fix:
            lines.append("\n✅ **Valores corrigidos para 0.**")
        else:
            lines.append("\n💡 Use `!fixkda sim` para corrigir os valores para 0.")

        await ctx.send("\n".join(lines))

    @bot.command(name="fixduracoes", aliases=["fixdurations", "corrigirduracoes"])
    async def cmd_fix_duracoes(ctx: commands.Context):
        if not is_admin(ctx.author.id):
            await ctx.message.delete()
            await ctx.send("❌ Apenas administradores.", delete_after=5)
            return

        fixed = fix_malformed_durations()
        if fixed:
            await ctx.send(f"✅ {fixed} duração(ões) corrigida(s) no banco (ex: `74:01` → `1:14:01`).")
        else:
            await ctx.send("✅ Nenhuma duração malformada encontrada.")

    @bot.command(name="renumerarpartida", aliases=["renamematch", "moveid"])
    async def cmd_renumerar_partida(ctx: commands.Context, old_id: int, new_id: int):
        if not is_admin(ctx.author.id):
            await ctx.message.delete()
            await ctx.send("❌ Apenas administradores.", delete_after=5)
            return
        try:
            renumber_league_match(old_id, new_id)
            await ctx.send(f"✅ Partida `#{old_id}` renumerada para `#{new_id}`.")
        except ValueError as e:
            await ctx.send(f"❌ {e}")

    @bot.command(name="fixsequencia", aliases=["fixseq", "resetsequencia"])
    async def cmd_fix_sequencia(ctx: commands.Context):
        if not is_admin(ctx.author.id):
            await ctx.message.delete()
            await ctx.send("❌ Apenas administradores.", delete_after=5)
            return

        max_id = fix_match_id_sequence()
        await ctx.send(f"✅ Sequência de IDs de partidas corrigida. Próxima partida será `#{max_id + 1}`.")

    @bot.command(name="diagtabela2", aliases=["debugtabela2", "diagtab2"])
    async def cmd_diag_tabela2(ctx: commands.Context):
        if not is_admin(ctx.author.id):
            await ctx.message.delete()
            await ctx.send("❌ Apenas administradores.", delete_after=5)
            return

        import traceback as _tb
        try:
            ranking = get_ranking_from_matches()
            await ctx.send(f"✅ `get_ranking_from_matches()` OK — {len(ranking)} jogadores sem erro.")
        except Exception as exc:
            tb_text = _tb.format_exc()
            short_tb = tb_text[-1800:] if len(tb_text) > 1800 else tb_text
            await ctx.send(
                f"❌ Erro em `get_ranking_from_matches()`:\n"
                f"```\n{type(exc).__name__}: {exc}\n```\n"
                f"```\n{short_tb}\n```"
            )

    _SUPER_ADMINS = ADMIN_IDS[:2]

    @bot.command(name="desligarbot", aliases=["botoff", "pausarbot"])
    async def cmd_bot_off(ctx: commands.Context):
        if ctx.author.id not in _SUPER_ADMINS:
            await ctx.message.delete()
            await ctx.send("❌ Apenas os dois primeiros administradores podem usar este comando.", delete_after=5)
            return
        if not BOT_STATE["enabled"]:
            await ctx.send("⚠️ O bot já está desativado.", delete_after=8)
            return
        BOT_STATE["enabled"] = False
        await ctx.message.delete()
        await ctx.send("🔴 Bot desativado. Comandos ignorados até `!ligarbot`.")

    @bot.command(name="ligarbot", aliases=["boton", "ativarbot"])
    async def cmd_bot_on(ctx: commands.Context):
        if ctx.author.id not in _SUPER_ADMINS:
            await ctx.message.delete()
            await ctx.send("❌ Apenas os dois primeiros administradores podem usar este comando.", delete_after=5)
            return
        if BOT_STATE["enabled"]:
            await ctx.send("⚠️ O bot já está ativo.", delete_after=8)
            return
        BOT_STATE["enabled"] = True
        await ctx.message.delete()
        await ctx.send("🟢 Bot reativado. Comandos voltando ao normal.")

    @bot.command(name="fechartemporada", aliases=["temporadaoff", "fimdetemporada"])
    async def cmd_season_off(ctx: commands.Context):
        if ctx.author.id not in _SUPER_ADMINS:
            await ctx.message.delete()
            await ctx.send("❌ Apenas os dois primeiros administradores podem usar este comando.", delete_after=5)
            return
        if not SEASON_STATE["active"]:
            await ctx.send("⚠️ A temporada já está encerrada.", delete_after=8)
            return
        SEASON_STATE["active"] = False
        await ctx.message.delete()
        await ctx.send("🏁 Temporada encerrada. Novas listas estão desativadas.")

    @bot.command(name="abrirtemporada", aliases=["temporadaon", "novaTemporada"])
    async def cmd_season_on(ctx: commands.Context):
        if ctx.author.id not in _SUPER_ADMINS:
            await ctx.message.delete()
            await ctx.send("❌ Apenas os dois primeiros administradores podem usar este comando.", delete_after=5)
            return
        if SEASON_STATE["active"]:
            await ctx.send("⚠️ A temporada já está ativa.", delete_after=8)
            return
        SEASON_STATE["active"] = True
        await ctx.message.delete()
        await ctx.send("🟢 Nova temporada iniciada. Listas liberadas.")

    @bot.command(name="exportar", aliases=["exportarpartidas", "exportdb"])
    async def cmd_exportar(ctx: commands.Context):
        if not is_admin(ctx.author.id):
            await ctx.message.delete()
            await ctx.send("❌ Apenas administradores.", delete_after=5)
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            await ctx.send("❌ `openpyxl` não está instalado no servidor. Adicione ao requirements.txt e faça redeploy.")
            return

        await ctx.message.delete()
        status_msg = await ctx.send("⏳ Gerando planilha...")

        from core.db.connection import get_connection

        with get_connection() as conn:
            rows = conn.execute("""
                SELECT
                    m.league_match_id,
                    m.external_match_id,
                    m.match_datetime,
                    m.winner_team,
                    m.duration,
                    m.score_radiant,
                    m.score_dire,
                    mp.slot,
                    mp.player_name,
                    COALESCE(p.display_name, mp.player_name) AS display_name,
                    mp.team,
                    mp.hero_name,
                    mp.kills,
                    mp.deaths,
                    mp.assists,
                    mp.networth,
                    CASE WHEN mp.team = m.winner_team THEN 'win' ELSE 'loss' END AS resultado,
                    m.created_at
                FROM matches m
                JOIN match_players mp ON mp.league_match_id = m.league_match_id
                LEFT JOIN players p ON p.discord_id = mp.discord_id
                ORDER BY m.league_match_id ASC, mp.slot ASC
            """).fetchall()

        HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
        HEADER_FONT = Font(color="FFFFFF", bold=True)
        ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
        WIN_FILL = PatternFill("solid", fgColor="C6EFCE")
        LOSS_FILL = PatternFill("solid", fgColor="FFC7CE")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Partidas"

        headers = [
            "Partida #", "Match ID Externo", "Data/Hora", "Time Vencedor",
            "Duração", "Placar Radiant", "Placar Dire",
            "Slot", "Nick no jogo", "Nome Discord", "Time",
            "Herói", "Kills", "Deaths", "Assists", "Networth", "Resultado", "Registrado em"
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(rows, start=2):
            values = list(row)
            for col_idx, value in enumerate(values, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

            resultado = row["resultado"]
            fill = WIN_FILL if resultado == "win" else LOSS_FILL
            result_col = headers.index("Resultado") + 1
            ws.cell(row=row_idx, column=result_col).fill = fill

            if row_idx % 2 == 0:
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                        cell.fill = ALT_FILL

        col_widths = [10, 18, 20, 15, 10, 14, 12, 6, 20, 20, 10, 20, 8, 8, 8, 12, 10, 20]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A2"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"partidas_{timestamp}.xlsx"

        await status_msg.edit(content=f"✅ {len(rows)} registros exportados.")
        await ctx.send(file=discord.File(buffer, filename=filename))

    @bot.command(name="cpi")
    async def cmd_cpi(ctx: commands.Context):
        from datetime import date
        await ctx.message.delete()

        hoje = date.today().strftime("%d/%m/%Y")

        nomes = []
        for aid in ADMIN_IDS[:2]:
            member = ctx.guild.get_member(aid) if ctx.guild else None
            nomes.append(member.display_name if member else str(aid))
        investigados = ", ".join(nomes)

        msg = (
            f"⚖️ **SUPREMO TRIBUNAL FEDERAL**\n"
            f"*Brasília, {hoje} — Plenário Virtual*\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"**INQUÉRITO Nº 0069/2025 — OPERAÇÃO FUMOS GATE**\n\n"
            f"O caso encontra-se sob condução dos Ministros Relatores:\n\n"
            f"**{investigados}**\n\n"
            f"*\"Após análise minuciosa dos autos, as evidências são absolutamente contundentes. "
            f"O investigado sabia, o investigado mandou, o investigado pagou "
            f"e ainda assim perdeu de Sniper. "
            f"A apuração segue em sigilo de Justiça.\"*\n\n"
            f"— **STF — 1ª Turma da Derrota**\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"🔨 *O silêncio pode ser usado contra você no ranking.*"
        )
        await ctx.send(msg)
